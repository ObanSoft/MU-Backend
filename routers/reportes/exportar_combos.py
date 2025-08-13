from flask import Blueprint, request, send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from models.Venta import Venta
from utils.auth_utils import token_required
from flask_cors import cross_origin

exportar_excel_combos_bp = Blueprint('exportar_excel_combos', __name__)

@exportar_excel_combos_bp.route('/exportar_combos', methods=['GET', 'OPTIONS'], strict_slashes=False)
@cross_origin(origin='http://localhost:3000', supports_credentials=True)
@token_required
def exportar_excel_combos():
    if request.method == 'OPTIONS':
        return '', 200

    # 🔹 Filtrar solo ventas tipo Combo
    ventas = Venta.query.filter_by(tipo_venta='Combo').all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Combos"

    encabezados = [
        'ID', 
        'Identificador Único', 
        'Nombre Producto', 
        'Precio', 
        'Fecha de Venta',
        'Vendido Por',
        'Método de Pago'
    ]
    ws.append(encabezados)

    for venta in ventas:
        ws.append([
            venta.id,
            venta.identificador_unico,
            venta.nombre_producto,
            float(venta.precio),
            venta.fecha_venta.strftime('%Y-%m-%d %H:%M:%S') if venta.fecha_venta else '',
            venta.vendido_por or '',
            venta.metodo_pago or ''
        ])

    # Estilos
    bold_font = Font(bold=True, color='FFFFFF')
    fill_color = PatternFill(start_color='E91E63', end_color='E91E63', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = fill_color
        cell.alignment = center_align
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            cell.alignment = Alignment(horizontal='left')
            if idx == 4:
                cell.number_format = '"$"#,##0.00'

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    table_range = f"A1:G{ws.max_row}"
    tabla = Table(displayName="TablaCombos", ref=table_range)
    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium10",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo_tabla
    ws.add_table(tabla)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_combos.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
